"""Carrier Flip Analysis for the Carrier Tender Optimization Dashboard.

Ports the standalone carrier-flip engine into the app so analysts can build the
flip report directly from the dashboard instead of running a separate script.

The report compares the ORIGINAL Dray SCAC on each container (from the GVT) with
the NEW SCAC assigned by a tender-optimization allocation, and looks up rate-card
rates for both carriers on the lane:

  - Old Rate = rate card rate for the ORIGINAL Dray SCAC on this lane
  - New Rate = rate card rate for the NEW SCAC (from the tender) on this lane
  - Savings  = Old Rate - New Rate (positive = cost reduction)

The engine functions (``run_carrier_flip_analysis`` and its helpers) are pure and
DataFrame-in / DataFrame-out so they can be unit tested without Streamlit. The
``show_carrier_flip_report`` function is the Streamlit UI wrapper that reuses the
already-loaded in-app GVT and Rate data (with optional file overrides) and offers
the multi-sheet Excel download.
"""

import io
import os
import re
import logging

import pandas as pd

logger = logging.getLogger(__name__)


KNOWN_CARRIERS = {
    'ATMI', 'ULSE', 'DMCQ', 'HDDR', 'HJBT', 'RKNE', 'KNIG',
    'RDXY', 'SONW', 'XPDR', 'PGLT', 'AOYV', 'FRQT', 'ARVY', 'AZGM'
}

# SCAC codes are 2-4 uppercase letters; we accept any 2-4 letter token as a
# potential carrier so that new carriers aren't silently dropped.
_SCAC_RE = re.compile(r'[A-Z]{2,4}')

# Regex matching a single container ID (4 letters + 7 digits, ISO 6346)
_CONTAINER_RE = re.compile(r'[A-Z]{3}[A-Z0-9]\d{6,7}')

# Collect carrier codes seen during parsing that are NOT in KNOWN_CARRIERS.
# Populated by parse_flip_info(), reset at the start of run_carrier_flip_analysis().
_unrecognized_carriers: set = set()


# ============================================================================
# Pure value helpers
# ============================================================================

def _parse_rate(value):
    """Convert a rate value to numeric float.

    Handles dollar-formatted strings ('$175.00', '$1,255.00'), plain strings
    ('175', '175.00'), and passthrough for values that are already numeric.
    Returns None for unparseable values.
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace('$', '').replace(',', '')
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _normalize_container(cid):
    """Normalize a container ID: uppercase, strip whitespace, non-alphanumeric chars, and trailing X markers."""
    if not isinstance(cid, str) or not cid.strip():
        return ''
    norm = re.sub(r'[^A-Z0-9]', '', cid.strip().upper())
    # Strip trailing 'X' markers used to flag cancelled/problem containers
    norm = norm.rstrip('X')
    return norm


def _normalize_container_short(cid):
    """Normalize container ID and strip ISO 6346 check digit (last digit) for fuzzy matching.

    Standard container IDs are 4 letters + 6 digits + 1 check digit = 11 chars.
    This returns the first 10 chars for matching containers that may differ
    only in their check digit.
    """
    norm = _normalize_container(cid)
    if len(norm) == 11 and re.match(r'^[A-Z]{3}[A-Z0-9]\d{7}$', norm):
        return norm[:10]
    return norm


# ============================================================================
# File reading / classification (used by the Streamlit UI)
# ============================================================================

def _read_any(file, name):
    """Load a CSV or Excel file (path or file-like) into a DataFrame using its name for the extension."""
    _, ext = os.path.splitext(name)
    ext = ext.lower()
    if ext in ('.xls', '.xlsx'):
        return pd.read_excel(file)
    elif ext == '.csv':
        return pd.read_csv(file)
    else:
        raise ValueError(f"Unsupported file format: {ext}")


def is_rate_card(file, name):
    """Check if an Excel file (path or file-like) is a Master Rate Card.

    Looks for any sheet whose name contains 'master' (e.g.
    'Master Sheet 2023 Award', 'Master Rate', 'US Dray Master - 3P', etc.).
    """
    _, ext = os.path.splitext(name)
    if ext.lower() not in ('.xls', '.xlsx'):
        return False
    try:
        xl = pd.ExcelFile(file)
        return any('master' in s.lower() for s in xl.sheet_names)
    except Exception:
        return False


def _find_master_sheet(xl):
    """Find the master rate sheet and its header row in an ExcelFile.

    Returns (sheet_name, header_row).  Skips sheets with 'Original' in
    the name so we prefer the current-award sheet over archived ones.
    Auto-detects the header row by scanning for a row that contains
    rate-card marker columns ('Lookup' + 'Base Rate', or 'SCAC' + 'Base Rate').
    """
    candidates = [
        s for s in xl.sheet_names
        if 'master' in s.lower() and 'original' not in s.lower()
    ]
    if not candidates:
        candidates = [s for s in xl.sheet_names if 'master' in s.lower()]
    if not candidates:
        return None, 0

    sheet_name = candidates[0]

    # Scan the first 10 rows to find the header that contains key columns
    preview = pd.read_excel(xl, sheet_name, header=None, nrows=10)
    header_row = 0
    for i, row in preview.iterrows():
        vals = [str(v).strip() for v in row.values if pd.notna(v)]
        # Accept headers with 'Lookup' + 'Base Rate' OR 'SCAC' + 'Base Rate'
        has_rate = 'Base Rate' in vals
        has_key = 'Lookup' in vals or 'SCAC' in vals
        if has_rate and has_key:
            header_row = i
            break

    return sheet_name, header_row


def load_rate_card(file):
    """Load the Master Rate Card: rates from the Master Sheet and port equivalences.

    Dynamically finds the correct sheet and header row so it works with
    different naming conventions (e.g. 'Master Sheet 2023 Award', 'Master Rate').
    Accepts a file path or a file-like object.
    """
    xl = pd.ExcelFile(file)
    sheet_name, header_row = _find_master_sheet(xl)
    if sheet_name is None:
        raise ValueError("No master rate sheet found in rate card")
    rates_df = pd.read_excel(xl, sheet_name, header=header_row)
    port_dup_df = None
    if 'Port Duplication' in xl.sheet_names:
        port_dup_df = pd.read_excel(xl, 'Port Duplication')
    return rates_df, port_dup_df


def classify_file(df):
    """Classify a dataframe as 'tender', 'constrained', or 'gvt' based on its columns."""
    cols_lower = [str(c).lower() for c in df.columns]
    # Strip emoji prefixes for robust matching (e.g. '📝 Description' → 'description')
    cols_stripped = [re.sub(r'^[^\w]+', '', c).strip() for c in cols_lower]

    # GVT: has Container + Dray SCAC — check first since it's unambiguous
    if 'container' in cols_lower and any('dray' in c and 'scac' in c for c in cols_lower):
        return 'gvt'

    has_scac = any('scac' in c for c in cols_lower)
    has_containers = 'container numbers' in cols_lower
    # Constrained indicator: any column containing 'description', 'priority', 'method',
    # or the exact 'constraint_description' — accepts emoji-prefixed columns
    has_constraint_desc = (
        'constraint_description' in cols_lower
        or any('description' in c for c in cols_stripped)
        or any('priority' in c for c in cols_stripped)
        or any('method' in c for c in cols_stripped)
    )

    # Constrained allocations: check BEFORE tender to avoid misclassification
    # when a constrained file also has a 'Carrier Flips' column
    if has_scac and has_containers and has_constraint_desc:
        return 'constrained'

    if 'carrier flips' in cols_lower:
        return 'tender'
    # Fallback: if it has Carrier and Container Numbers, it's tender
    if 'carrier' in cols_lower and has_containers:
        return 'tender'
    return 'unknown'


# ============================================================================
# Rate-card lookup
# ============================================================================

def build_rate_lookup(rates_df, port_dup_df=None):
    """Build a lookup dict: (SCAC + Lane) -> Base Rate.

    The rate card 'Lookup' column is SCAC+Port+FC (e.g. 'ATMIUSBWIBWI4').
    The tender 'Lane' column is Port+FC (e.g. 'USBWIBWI4').
    So the key is  Dray_SCAC + Lane.

    Also tries alternative rate column names: 'Base Rate', 'Rate', 'Contract Rate',
    'Award Rate', 'Awarded Rate' — uses the first one found.

    Port equivalences from the Port Duplication sheet are expanded so that
    lookups work regardless of which port alias is used.
    """
    if rates_df is None:
        return {}

    # Find the rate column — try multiple common names
    rate_col = None
    for candidate in ['Base Rate', 'Rate', 'Contract Rate', 'Award Rate', 'Awarded Rate', 'base rate']:
        if candidate in rates_df.columns:
            rate_col = candidate
            break
    # Case-insensitive fallback
    if rate_col is None:
        for col in rates_df.columns:
            if 'rate' in col.lower() and 'total' not in col.lower():
                rate_col = col
                break

    has_lookup = 'Lookup' in rates_df.columns
    has_components = 'SCAC' in rates_df.columns and 'Port' in rates_df.columns and 'FC' in rates_df.columns

    if not has_lookup and not has_components:
        logger.warning("Rate card missing 'Lookup' and 'SCAC'/'Port'/'FC' columns. Columns: %s",
                       list(rates_df.columns))
        return {}
    if rate_col is None:
        logger.warning("Rate card missing rate column. Columns: %s", list(rates_df.columns))
        return {}

    rate_map = {}

    # Build keys from 'Lookup' column if present
    if has_lookup:
        for _, row in rates_df.iterrows():
            lookup = row.get('Lookup')
            rate = row.get(rate_col)
            if pd.notna(lookup) and pd.notna(rate):
                key = str(lookup).strip().upper()
                rate_map[key] = rate

    # Build keys from separate SCAC+Port+FC columns
    if has_components:
        added = 0
        for _, row in rates_df.iterrows():
            scac_val = row.get('SCAC')
            port_val = row.get('Port')
            fc_val = row.get('FC')
            rate = row.get(rate_col)
            if all(pd.notna(x) for x in [scac_val, port_val, fc_val, rate]):
                key = str(scac_val).strip().upper() + str(port_val).strip().upper() + str(fc_val).strip().upper()
                if key not in rate_map:
                    rate_map[key] = rate
                    added += 1
        if added:
            source = "primary" if not has_lookup else "extra"
            logger.info("Rate card: %d %s entries from SCAC+Port+FC columns", added, source)

    logger.info("Rate card: %d entries loaded from column '%s'", len(rate_map), rate_col)

    # Expand with port equivalences so both aliases resolve
    if port_dup_df is not None and 'Port' in port_dup_df.columns and 'Equivalent' in port_dup_df.columns:
        port_equiv = {}
        for _, row in port_dup_df.iterrows():
            p, eq = str(row['Port']).strip().upper(), str(row['Equivalent']).strip().upper()
            port_equiv[p] = eq
            port_equiv[eq] = p

        # For every existing rate entry, add the equivalent-port version
        extras = {}
        for key, rate in rate_map.items():
            scac_match = re.match(r'^([A-Z]{2,4})', key)
            if scac_match:
                scac = scac_match.group(1)
                rest = key[len(scac):]
                # Try each port length (4-6 chars) to find a match in port_equiv
                for port_len in range(4, 7):
                    if len(rest) >= port_len:
                        port = rest[:port_len]
                        fc = rest[port_len:]
                        if port in port_equiv:
                            alt_key = scac + port_equiv[port] + fc
                            if alt_key not in rate_map:
                                extras[alt_key] = rate
                            break
        rate_map.update(extras)
        if extras:
            logger.info("Port equivalences added %d extra rate entries", len(extras))

    return rate_map


def _lookup_rate_from_map(scac, lane, fc, rate_map):
    """Look up a carrier rate from the rate map using tiered strategies.

    Returns (rate, method) where method is 'exact', 'port', 'fc', or None.
    Only returns a rate when the match is unambiguous (single match).
    All key components are trimmed and uppercased for reliable matching.
    """
    if not scac or not str(scac).strip():
        return None, None
    scac = str(scac).strip().upper()

    # Strategy 1: exact SCAC + Lane (Lane = Port + FC)
    if lane and str(lane).strip():
        key = scac + str(lane).strip().upper()
        if key in rate_map:
            return rate_map[key], 'exact'

    # Strategy 2: SCAC + port prefix (without FC suffix)
    # Only use if exactly ONE rate_map entry matches to avoid ambiguity
    if lane and len(str(lane).strip()) >= 5:
        lane_str = str(lane).strip().upper()
        port = lane_str[:5]
        key_port_only = scac + port
        matches = [(rk, rv) for rk, rv in rate_map.items() if rk.startswith(key_port_only)]
        if len(matches) == 1:
            return matches[0][1], 'port'

    # Strategy 3: SCAC + FC suffix
    # Only use if exactly ONE rate_map entry matches to avoid cross-port contamination
    if fc and str(fc).strip():
        fc_str = str(fc).strip().upper()
        matches = [(rk, rv) for rk, rv in rate_map.items()
                   if rk.startswith(scac) and rk.endswith(fc_str)]
        if len(matches) == 1:
            return matches[0][1], 'fc'

    return None, None


def _find_dray_scac_col(df):
    """Find the Dray SCAC column in a GVT dataframe.

    Matches columns containing both 'dray' and 'scac' (case-insensitive).
    Returns the actual column name, or None if not found.
    """
    for col in df.columns:
        low = col.lower()
        if 'dray' in low and 'scac' in low:
            return col
    return None


def _strip_facility_suffix(fac):
    """Strip suffixes like '-S', '-NS', '-IXD' from facility codes to get bare FC."""
    if not isinstance(fac, str) or not fac.strip():
        return ''
    return re.sub(r'-[A-Z0-9]+$', '', fac.strip().upper())


def lookup_old_carrier_rate(gvt_merged, rate_map, dray_scac_col='Dray SCAC(FL)'):
    """Add 'Old Rate', 'Old Rate Method', and 'Old Rate Reason' columns to GVT.

    Looks up old carrier (Dray SCAC) + Lane in rate_map using strict
    tiered strategies that only return a rate when the match is unambiguous.

    When Lane is missing (container not in tender), constructs a lookup
    lane from GVT's own Discharged Port + Facility columns:
      Lane = 'US' + Port + stripped_Facility  (e.g. 'USLAX' + 'LAX9' = 'USLAXLAX9')

    When no rate is found, 'Old Rate Reason' explains why.
    """
    methods = []
    reasons = []

    # Pre-compute set of SCACs that exist in the rate map
    rate_map_scacs = set()
    for key in rate_map:
        m = re.match(r'^([A-Z]{2,4})', key)
        if m:
            rate_map_scacs.add(m.group(1))

    def _get_rate(row):
        scac = row.get(dray_scac_col)
        lane = row.get('Lane') if pd.notna(row.get('Lane')) else None
        fc = row.get('FC') if pd.notna(row.get('FC')) else None

        # Fallback: construct lane from GVT's own Port + Facility
        if not lane:
            port = row.get('Discharged Port')
            fac = row.get('Facility')
            if pd.notna(port) and str(port).strip():
                port_str = 'US' + str(port).strip().upper()
                fc_str = _strip_facility_suffix(fac) if pd.notna(fac) else ''
                lane = port_str + fc_str
                if not fc and fc_str:
                    fc = fc_str

        if pd.isna(scac) or not str(scac).strip():
            methods.append(None)
            reasons.append('No Dray SCAC assigned')
            return None

        scac_upper = str(scac).strip().upper()
        rate, method = _lookup_rate_from_map(scac_upper, lane, fc, rate_map)
        methods.append(method)

        if rate is not None:
            reasons.append(None)
        elif scac_upper not in rate_map_scacs:
            reasons.append(f'SCAC {scac_upper} not in rate card')
        else:
            # SCAC exists but no rate for this specific lane/FC combo
            fc_display = fc or _strip_facility_suffix(row.get('Facility')) if pd.notna(row.get('Facility')) else ''
            port_display = ''
            if lane and len(str(lane).strip()) >= 5:
                port_display = str(lane).strip().upper()[:5]
            elif pd.notna(row.get('Discharged Port')):
                port_display = 'US' + str(row['Discharged Port']).strip().upper()
            reasons.append(f'No rate for {scac_upper} at {port_display}+{fc_display}')
        return rate

    gvt_merged['Old Rate'] = gvt_merged.apply(_get_rate, axis=1)
    gvt_merged['Old Rate Method'] = methods
    gvt_merged['Old Rate Reason'] = reasons

    found = gvt_merged['Old Rate'].notna().sum()
    total = len(gvt_merged)
    logger.info("Old Rate lookup: %d/%d matched (exact: %d, port: %d, fc: %d)",
                found, total, methods.count('exact'), methods.count('port'), methods.count('fc'))

    return gvt_merged


# ============================================================================
# Tender allocation parsing
# ============================================================================

def parse_flip_info(text):
    """Parse the 'Carrier Flips' column into structured data.

    Accepts ANY 2-4 letter uppercase SCAC code (not just KNOWN_CARRIERS).
    Unrecognized codes are still included in the output but also tracked
    in ``_unrecognized_carriers`` so the caller can warn the user.

    Returns a Series: [Now, From, Lost, From Count, Lost Count].
    """
    if not isinstance(text, str):
        return pd.Series([0, "", "", 0, 0])

    now_match = re.search(r"Now\s+(\d+)", text)
    now_count = int(now_match.group(1)) if now_match else 0

    from_matches = re.findall(r"([A-Z]{2,4})\s+\(\+(\d+)\)", text)
    from_entries = []
    from_count = 0
    for c, n in from_matches:
        # Skip the literal word "Now" which the regex could match
        if c == 'Now':
            continue
        from_entries.append(f"{c} (+{n})")
        from_count += int(n)
        if c not in KNOWN_CARRIERS:
            _unrecognized_carriers.add(c)

    to_matches = re.findall(r"([A-Z]{2,4})\s+\(\-(\d+)\)", text)
    to_entries = []
    lost_count = 0
    for c, n in to_matches:
        if c == 'Now':
            continue
        to_entries.append(f"{c} (-{n})")
        lost_count += int(n)
        if c not in KNOWN_CARRIERS:
            _unrecognized_carriers.add(c)

    return pd.Series([now_count, ", ".join(from_entries), ", ".join(to_entries), from_count, lost_count])


def _split_container_numbers(text):
    """Split a 'Container Numbers' cell into individual container IDs.

    Handles comma, semicolon, pipe, newline, and whitespace separators.
    """
    if not isinstance(text, str) or not text.strip():
        return []

    tokens = re.split(r'[,;|\n\r]+', text)
    tokens = [t.strip() for t in tokens if t.strip()]
    return tokens


def create_container_carrier_mapping(df_tender):
    """Create a mapping of container -> new carrier from tender optimization data."""
    empty = pd.DataFrame(
        columns=['Container', 'NEW SCAC', 'Discharged Port', 'Category', 'Week Number', 'New Rate', 'Lane']
    )
    if df_tender is None or df_tender.empty or 'Container Numbers' not in df_tender.columns:
        return empty

    # Support both 'NEW SCAC' (unconstrained) and 'Carrier' column names
    carrier_col = None
    for col in ['NEW SCAC', 'Carrier', 'SCAC']:
        if col in df_tender.columns:
            carrier_col = col
            break
    if carrier_col is None:
        return empty

    # Find the rate column — try multiple common names from tender optimization files
    rate_col = None
    for candidate in ['Base Rate', 'Rate', 'Contract Rate', 'Award Rate', 'Awarded Rate', 'New Rate']:
        if candidate in df_tender.columns:
            rate_col = candidate
            break
    if rate_col is None:
        for col in df_tender.columns:
            if 'rate' in col.lower() and 'total' not in col.lower():
                rate_col = col
                break

    df_with_containers = df_tender[df_tender['Container Numbers'].notna()].copy()
    df_with_containers['Container'] = df_with_containers['Container Numbers'].apply(_split_container_numbers)
    df_exploded = df_with_containers.explode('Container')
    df_exploded['Container'] = df_exploded['Container'].apply(_normalize_container)
    # Drop rows where container ID is empty after normalization
    df_exploded = df_exploded[df_exploded['Container'].astype(bool)]
    # Drop rows where carrier is NaN — these are unallocated containers
    total_before = len(df_exploded)
    df_exploded = df_exploded[df_exploded[carrier_col].notna()]
    skipped = total_before - len(df_exploded)
    if skipped > 0:
        logger.info("Skipped %d containers with no carrier assignment", skipped)

    select_cols = ['Container', carrier_col]
    for opt in ['Discharged Port', 'Category', 'Week Number']:
        if opt in df_exploded.columns:
            select_cols.append(opt)
    if rate_col and rate_col in df_exploded.columns:
        select_cols.append(rate_col)
    if 'Lane' in df_exploded.columns:
        select_cols.append('Lane')
    if 'FC' in df_exploded.columns:
        select_cols.append('FC')

    result = df_exploded[select_cols].copy().rename(columns={carrier_col: 'NEW SCAC'})
    # Rename the rate column to 'New Rate' for clarity
    if rate_col and rate_col in result.columns and rate_col != 'New Rate':
        result = result.rename(columns={rate_col: 'New Rate'})

    # Derive FC from Lane if not already present (Lane = Port(5) + FC)
    if 'FC' not in result.columns and 'Lane' in result.columns:
        result['FC'] = result['Lane'].apply(
            lambda x: str(x).strip().upper()[5:] if pd.notna(x) and len(str(x).strip()) > 5 else None
        )

    # Convert dollar-formatted rates to numeric (e.g. '$175.00' -> 175.0)
    if 'New Rate' in result.columns:
        result['New Rate'] = result['New Rate'].apply(_parse_rate)

    return result


def _find_gvt_container_col(gvt_df):
    """Find the container ID column in a GVT dataframe.

    Checks for common names: 'Container', 'Container ID', 'container_id', etc.
    Returns the actual column name.
    """
    for col in gvt_df.columns:
        if col.lower().replace(' ', '').replace('_', '') in ('container', 'containerid'):
            return col
    for col in gvt_df.columns:
        if 'container' in col.lower():
            return col
    return None


# ============================================================================
# GVT merge + savings
# ============================================================================

def merge_gvt_with_carrier_flips(gvt_df, container_mapping, rate_map=None):
    """Merge GVT data with new carrier assignments, new rate, lane, old rate, and savings.

    Uses a left join so every GVT row is kept; tender containers not in GVT do
    not create new rows. Adds check-digit-stripped fallback matching for
    unmatched containers.
    """
    gvt_container_col = _find_gvt_container_col(gvt_df)
    if gvt_container_col is None:
        logger.warning("No container column found in GVT data. Columns: %s", list(gvt_df.columns))
        return gvt_df

    # Normalize container IDs in GVT so they match the mapping
    gvt_df = gvt_df.copy()
    gvt_df['_container_key'] = gvt_df[gvt_container_col].apply(_normalize_container)
    container_mapping = container_mapping.copy()
    container_mapping['_container_key'] = container_mapping['Container'].apply(_normalize_container)

    merge_cols = ['_container_key', 'NEW SCAC']
    for opt_col in ['New Rate', 'Lane', 'Week Number', 'FC']:
        if opt_col in container_mapping.columns:
            merge_cols.append(opt_col)

    # LEFT join: keep all GVT rows intact, add new SCAC/rate where matched
    merged = gvt_df.merge(
        container_mapping[merge_cols],
        on='_container_key',
        how='left',
        indicator=True
    )

    # Check-digit fallback for unmatched GVT containers
    unmatched_gvt_mask = (merged['_merge'] == 'left_only')
    unmatched_gvt_count = unmatched_gvt_mask.sum()
    if unmatched_gvt_count > 0:
        mapping_short = container_mapping.copy()
        mapping_short['_short_key'] = mapping_short['Container'].apply(_normalize_container_short)
        short_lookup = mapping_short.drop_duplicates(subset=['_short_key'], keep='first')
        short_map = short_lookup.set_index('_short_key')[merge_cols[1:]].to_dict('index')

        recovered = 0
        for idx in merged.index[unmatched_gvt_mask]:
            full_key = merged.at[idx, '_container_key']
            short_key = _normalize_container_short(full_key) if full_key else ''
            if short_key and short_key in short_map:
                for col_name in merge_cols[1:]:
                    merged.at[idx, col_name] = short_map[short_key].get(col_name)
                merged.at[idx, '_merge'] = 'both'
                recovered += 1
        if recovered > 0:
            logger.info("Check-digit fallback: recovered %d/%d previously unmatched GVT containers",
                        recovered, unmatched_gvt_count)

    merged['Match Status'] = merged['_merge'].map({
        'both': 'Matched',
        'left_only': 'No New Assignment'
    })
    merged.drop(columns=['_container_key', '_merge'], inplace=True)

    status_counts = merged['Match Status'].value_counts()
    logger.info("GVT merge: %d total rows (matched: %d, no new assignment: %d, container col: '%s')",
                len(merged), status_counts.get('Matched', 0),
                status_counts.get('No New Assignment', 0), gvt_container_col)

    total = len(merged)
    # If rate card was provided, look up rates from the rate card for both carriers.
    if rate_map:
        dray_scac_col = _find_dray_scac_col(merged) or 'Dray SCAC(FL)'
        merged = lookup_old_carrier_rate(merged, rate_map, dray_scac_col)

        new_methods = []

        def _get_new_rate(row):
            scac = row.get('NEW SCAC')
            lane = row.get('Lane') if pd.notna(row.get('Lane')) else None
            fc = row.get('FC') if pd.notna(row.get('FC')) else None
            if pd.isna(scac) or not str(scac).strip():
                new_methods.append(None)
                return None
            rate, method = _lookup_rate_from_map(str(scac).strip(), lane, fc, rate_map)
            new_methods.append(method)
            return rate

        merged['New Rate'] = merged.apply(_get_new_rate, axis=1)
        merged['New Rate Method'] = new_methods

        rc_found = merged['New Rate'].notna().sum()
        logger.info("New Rate lookup (rate card for NEW SCAC): %d/%d matched", rc_found, total)

        def _calc_savings(row):
            old = row.get('Old Rate')
            new = row.get('New Rate')
            if pd.notna(old) and pd.notna(new):
                try:
                    return float(old) - float(new)
                except (ValueError, TypeError):
                    return None
            return None
        merged['Savings'] = merged.apply(_calc_savings, axis=1)

    # Reorder: put key analysis columns right after Dray SCAC column
    cols = list(merged.columns)
    dray_scac_col = _find_dray_scac_col(merged)
    if dray_scac_col and 'NEW SCAC' in cols:
        insert_after = []
        for c in ['NEW SCAC', 'Lane', 'New Rate', 'New Rate Method',
                  'Old Rate', 'Old Rate Reason', 'Old Rate Method', 'Savings', 'Match Status']:
            if c in cols:
                cols.remove(c)
                insert_after.append(c)
        if dray_scac_col in cols:
            dray_idx = cols.index(dray_scac_col)
            for i, c in enumerate(insert_after):
                cols.insert(dray_idx + 1 + i, c)
        merged = merged[cols]
    return merged


def build_match_diagnostics(gvt_df, combined_mapping, gvt_merged, rate_map):
    """Return a list of human-readable diagnostic lines explaining match coverage."""
    lines = []

    gvt_container_col = _find_gvt_container_col(gvt_df)
    if gvt_container_col is None:
        return ["No container column found in GVT data."]

    gvt_keys = set(gvt_df[gvt_container_col].apply(_normalize_container).dropna()) - {''}
    tender_keys = set(combined_mapping['Container'].apply(_normalize_container).dropna()) - {''}

    lines.append(f"Unique containers — GVT: {len(gvt_keys)}, Tender: {len(tender_keys)}")
    lines.append(f"In both: {len(gvt_keys & tender_keys)}")
    lines.append(f"In GVT only (no flip): {len(gvt_keys - tender_keys)}")
    lines.append(f"In tender only (not in GVT): {len(tender_keys - gvt_keys)}")

    if 'New Rate' in gvt_merged.columns and 'NEW SCAC' in gvt_merged.columns:
        flipped = gvt_merged[gvt_merged['NEW SCAC'].notna()]
        no_new_rate = flipped[flipped['New Rate'].isna()]
        lines.append(f"Flipped containers missing New Rate (rate-card gap): {len(no_new_rate)}")
        if len(no_new_rate) > 0 and 'Lane' in no_new_rate.columns:
            gaps = no_new_rate.groupby(['NEW SCAC', 'Lane']).size().sort_values(ascending=False)
            for (scac, lane), n in gaps.head(10).items():
                lines.append(f"  no rate: {scac} @ {lane}: {n} containers")

    if 'Old Rate Reason' in gvt_merged.columns:
        reasons = gvt_merged['Old Rate Reason'].dropna()
        for reason, n in reasons.value_counts().head(10).items():
            lines.append(f"  Old Rate not found ({n}): {reason}")

    if rate_map and 'NEW SCAC' in gvt_merged.columns:
        rate_map_scacs = set()
        for key in rate_map:
            m = re.match(r'^([A-Z]{2,4})', key)
            if m:
                rate_map_scacs.add(m.group(1))
        new_scacs = set(gvt_merged['NEW SCAC'].dropna().astype(str).str.strip().str.upper())
        missing_scacs = new_scacs - rate_map_scacs
        if missing_scacs:
            lines.append(f"NEW SCACs not in rate card at all: {sorted(missing_scacs)}")

    return lines


# ============================================================================
# Orchestration (pure — returns DataFrames, no I/O)
# ============================================================================

def run_carrier_flip_analysis(tender_dfs=None, constrained_dfs=None,
                              gvt_df=None, rates_df=None, port_dup_df=None):
    """Run the full carrier-flip analysis from already-loaded DataFrames.

    Args:
        tender_dfs: list of unconstrained tender allocation DataFrames (or None)
        constrained_dfs: list of constrained allocation DataFrames (or None)
        gvt_df: per-container GVT DataFrame (with Container + Dray SCAC) or None
        rates_df: rate-card DataFrame (with Lookup or SCAC/Port/FC + a rate column)
        port_dup_df: optional port-duplication DataFrame for alias expansion

    Returns a dict with keys:
        summary, unconstrained, constrained, gvt_merged (DataFrames or None),
        stats (dict), diagnostics (list[str]), unrecognized_carriers (sorted list).
    """
    # Reset the module-level unrecognized-carrier accumulator for this run
    _unrecognized_carriers.clear()

    tender_dfs = [d for d in (tender_dfs or []) if d is not None and not d.empty]
    constrained_dfs = [d for d in (constrained_dfs or []) if d is not None and not d.empty]

    result = {
        'summary': None,
        'unconstrained': None,
        'constrained': None,
        'gvt_merged': None,
        'stats': {},
        'diagnostics': [],
        'unrecognized_carriers': [],
    }

    if not tender_dfs and not constrained_dfs:
        result['diagnostics'].append("No tender optimization allocations provided.")
        return result

    # ---- Combine tender (unconstrained) files ----
    df_combined = None
    df_final = pd.DataFrame()
    if tender_dfs:
        df_combined = pd.concat(tender_dfs, ignore_index=True)

        if 'Carrier Flips' in df_combined.columns:
            has_flips = (df_combined['Carrier Flips'] != 'No Flip').any()
            if has_flips:
                new_columns = df_combined['Carrier Flips'].apply(parse_flip_info)
                new_columns.columns = ['Now', 'From', 'Lost', 'From Count', 'Lost Count']
                df_combined = pd.concat([df_combined, new_columns], axis=1)

        carrier_col = 'NEW SCAC' if 'NEW SCAC' in df_combined.columns else (
            'Carrier' if 'Carrier' in df_combined.columns else None)
        sort_cols = [c for c in [carrier_col, 'Discharged Port', 'Category', 'Week Number']
                     if c and c in df_combined.columns]
        if sort_cols:
            df_combined = df_combined.sort_values(by=sort_cols)

        if 'Base Rate' in df_combined.columns:
            df_combined = df_combined.rename(columns={'Base Rate': 'New Rate'})
        if 'New Rate' in df_combined.columns:
            df_combined['New Rate'] = df_combined['New Rate'].apply(_parse_rate)

        final_columns = ['Discharged Port', 'Category', carrier_col, 'Week Number',
                         'Lane', 'FC', 'Facility', 'Container Numbers', 'Container Count',
                         'New Rate', 'Total Cost', 'Carrier Flips']
        if 'Now' in df_combined.columns:
            final_columns.extend(['Now', 'From', 'Lost', 'From Count', 'Lost Count'])
        df_final = df_combined[[c for c in final_columns if c and c in df_combined.columns]].copy()
        if carrier_col and carrier_col != 'Carrier':
            df_final = df_final.rename(columns={carrier_col: 'Carrier'})

    # ---- Combine constrained files ----
    df_constrained = None
    if constrained_dfs:
        df_constrained = pd.concat(constrained_dfs, ignore_index=True)
        if 'Base Rate' in df_constrained.columns:
            df_constrained = df_constrained.rename(columns={'Base Rate': 'New Rate'})
        if 'New Rate' in df_constrained.columns:
            df_constrained['New Rate'] = df_constrained['New Rate'].apply(_parse_rate)

    # ---- Build container -> new carrier mapping ----
    # A container can appear in BOTH the unconstrained allocation (where the
    # optimizer is free to pick any carrier) and the constrained allocation (where
    # a user constraint LOCKS it to a specific carrier). The constraint must win:
    # tag each part with a source priority (constrained = 0, unconstrained = 1) and
    # break the per-container dedup on that priority FIRST, so the carrier a
    # constraint pinned the container to is the one the flip report reports.
    mapping_parts = []
    if df_constrained is not None:
        m = create_container_carrier_mapping(df_constrained)
        m['_source_priority'] = 0  # constrained assignment — authoritative
        mapping_parts.append(m)
    if df_combined is not None:
        m = create_container_carrier_mapping(df_combined)
        m['_source_priority'] = 1  # unconstrained optimizer — yields to a constraint
        mapping_parts.append(m)
    combined_mapping = pd.concat(mapping_parts, ignore_index=True) if mapping_parts else None
    if combined_mapping is not None and not combined_mapping.empty:
        combined_mapping = combined_mapping.sort_values(
            ['_source_priority', 'NEW SCAC'], na_position='last'
        ).drop_duplicates(subset=['Container'], keep='first')
        combined_mapping = combined_mapping.drop(columns=['_source_priority'])

    result['summary'] = df_final if not df_final.empty else None
    result['unconstrained'] = df_combined
    result['constrained'] = df_constrained

    # ---- Merge against GVT + rate card ----
    if gvt_df is not None and combined_mapping is not None and not combined_mapping.empty:
        rate_map = build_rate_lookup(rates_df, port_dup_df) if rates_df is not None else None
        gvt_merged = merge_gvt_with_carrier_flips(gvt_df, combined_mapping, rate_map)
        result['gvt_merged'] = gvt_merged

        stats = {}
        stats['gvt_rows'] = len(gvt_merged)
        if 'NEW SCAC' in gvt_merged.columns:
            stats['matched'] = int(gvt_merged['NEW SCAC'].notna().sum())
        if 'Old Rate' in gvt_merged.columns:
            stats['old_rate_found'] = int(gvt_merged['Old Rate'].notna().sum())
        if 'New Rate' in gvt_merged.columns:
            stats['new_rate_found'] = int(gvt_merged['New Rate'].notna().sum())
        if 'Savings' in gvt_merged.columns:
            stats['total_savings'] = float(gvt_merged['Savings'].sum())
        result['stats'] = stats

        result['diagnostics'] = build_match_diagnostics(gvt_df, combined_mapping, gvt_merged, rate_map)

    result['unrecognized_carriers'] = sorted(_unrecognized_carriers)
    return result


# ============================================================================
# Excel export
# ============================================================================

def _sanitize_df_for_excel(df):
    """Strip timezone info from datetime columns so openpyxl can write them."""
    df = df.copy()
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            try:
                df[col] = df[col].dt.tz_localize(None) if df[col].dt.tz else df[col]
            except (AttributeError, TypeError):
                pass
    return df


def build_flip_report_excel(results):
    """Build the multi-sheet flip-report workbook from run_carrier_flip_analysis() output.

    Returns the workbook as bytes (for st.download_button), or None if there is
    nothing to write.
    """
    from openpyxl.styles import PatternFill

    sheets = []
    if results.get('summary') is not None:
        sheets.append(('Carrier Flips Summary', results['summary']))
    if results.get('unconstrained') is not None:
        sheets.append(('Unconstrained Allocations', results['unconstrained']))
    if results.get('constrained') is not None:
        sheets.append(('Constrained Allocations', results['constrained']))
    if results.get('gvt_merged') is not None:
        sheets.append(('GVT with New SCAC', results['gvt_merged']))

    if not sheets:
        return None

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        for name, df in sheets:
            _sanitize_df_for_excel(df).to_excel(writer, sheet_name=name, index=False)

        # Highlight flipped carriers on the GVT sheet
        if 'GVT with New SCAC' in writer.sheets:
            yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            ws = writer.sheets['GVT with New SCAC']
            header_row = [cell.value for cell in ws[1]]
            dray_col = None
            for idx, h in enumerate(header_row):
                if h and 'dray' in str(h).lower() and 'scac' in str(h).lower():
                    dray_col = idx + 1
                    break
            new_col = header_row.index('NEW SCAC') + 1 if 'NEW SCAC' in header_row else None
            if dray_col and new_col:
                for row_num in range(2, ws.max_row + 1):
                    dray_val = ws.cell(row=row_num, column=dray_col).value
                    new_val = ws.cell(row=row_num, column=new_col).value
                    if new_val and new_val != '' and dray_val != new_val:
                        ws.cell(row=row_num, column=new_col).fill = yellow

    buf.seek(0)
    return buf.getvalue()


# ============================================================================
# Streamlit UI
# ============================================================================

def show_carrier_flip_report(in_app_gvt=None, in_app_rate=None):
    """Render the Carrier Flip Analysis section in the dashboard.

    Builds the flip report entirely from data already in the app: the allocation
    the Detailed Analysis Table just computed (stashed in session state) plus the
    already-loaded GVT (per-container) and Rate data. No file upload required.

    Args:
        in_app_gvt: the per-container GVTdata DataFrame already loaded in the app.
        in_app_rate: the processed Ratedata DataFrame already loaded in the app.
    """
    import streamlit as st
    from ..core.config_styling import section_header

    section_header("🔁 Carrier Flip Analysis")

    # Allocation(s) computed by the Detailed Analysis Table above. These are the
    # exact frames the table renders, so we consume them directly — no re-export.
    strategy = st.session_state.get('flip_source_strategy')
    in_app_unconstrained = st.session_state.get('flip_source_unconstrained')
    in_app_constrained = st.session_state.get('flip_source_constrained')

    with st.expander("ℹ️ What this does", expanded=False):
        st.markdown(
            "Compares each container's **original** Dray SCAC (from the GVT) against the "
            "**new** SCAC assigned by the tender allocation, then looks up rate-card rates "
            "for both carriers on the lane:\n\n"
            "- **Old Rate** — rate card rate for the original Dray SCAC\n"
            "- **New Rate** — rate card rate for the new SCAC\n"
            "- **Savings** — Old Rate − New Rate (positive = cost reduction)\n\n"
            "This runs automatically on the scenario selected in the **Detailed Analysis "
            "Table** above, reusing the GVT and Rate data already loaded in the app."
        )

    tender_dfs, constrained_dfs = [], []
    rates_df, port_dup_df = None, None
    gvt_df = in_app_gvt

    # ---- Use the allocation the Detailed Analysis Table produced ----
    if isinstance(in_app_unconstrained, pd.DataFrame) and not in_app_unconstrained.empty:
        tender_dfs.append(in_app_unconstrained)
    if isinstance(in_app_constrained, pd.DataFrame) and not in_app_constrained.empty:
        constrained_dfs.append(in_app_constrained)
    if (tender_dfs or constrained_dfs) and strategy:
        st.caption(f"Analyzing the **{strategy}** allocation from the Detailed Analysis Table above.")

    # ---- Rate card: reuse the in-app rate data ----
    if rates_df is None and in_app_rate is not None:
        rates_df = in_app_rate

    if not tender_dfs and not constrained_dfs:
        st.info(
            "Run the Detailed Analysis Table above to generate an allocation — "
            "the flip report will build from it automatically."
        )
        return

    with st.spinner("🔁 Building carrier flip report..."):
        results = run_carrier_flip_analysis(
            tender_dfs=tender_dfs,
            constrained_dfs=constrained_dfs,
            gvt_df=gvt_df,
            rates_df=rates_df,
            port_dup_df=port_dup_df,
        )

    # ---- Summary metrics ----
    stats = results.get('stats', {})
    if stats:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("📦 GVT Rows", f"{stats.get('gvt_rows', 0):,}")
        c2.metric("🔗 Matched", f"{stats.get('matched', 0):,}")
        c3.metric("💵 Rates Found (Old/New)",
                  f"{stats.get('old_rate_found', 0):,} / {stats.get('new_rate_found', 0):,}")
        c4.metric("💰 Total Savings", f"${stats.get('total_savings', 0):,.2f}")
    elif gvt_df is None:
        st.info("No GVT data available — showing allocation summary only. "
                "Load GVT data to compute Old/New rates and savings.")

    if results.get('unrecognized_carriers'):
        st.warning("⚠️ Unrecognized SCAC code(s) seen in Carrier Flips: "
                   f"{', '.join(results['unrecognized_carriers'])}")

    # ---- Result tables ----
    if results.get('gvt_merged') is not None:
        st.markdown("**GVT with New SCAC**")
        st.dataframe(results['gvt_merged'], use_container_width=True, hide_index=True)
    elif results.get('summary') is not None:
        st.markdown("**Carrier Flips Summary**")
        st.dataframe(results['summary'], use_container_width=True, hide_index=True)

    if results.get('diagnostics'):
        with st.expander("🔬 Match diagnostics", expanded=False):
            for line in results['diagnostics']:
                st.text(line)

    # ---- Excel download ----
    excel_bytes = build_flip_report_excel(results)
    if excel_bytes is not None:
        st.download_button(
            label="📥 Download Carrier Flip Analysis (Excel)",
            data=excel_bytes,
            file_name="carrier_flip_analysis.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="flip_download",
        )
